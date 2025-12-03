# openpyxl para arquivos Excel xlsx, xlsm, xltx e xltm (instalação)
# Com essa biblioteca será possível ler e escrever dados em células
# específicas, formatar células, inserir gráficos,
# criar fórmulas, adicionar imagens e outros elementos gráficos às suas
# planilhas. Ela é útil para automatizar tarefas envolvendo planilhas do
# Excel, como a criação de relatórios e análise de dados e/ou facilitando a
# manipulação de grandes quantidades de informações.
# Instalação necessária: pip install openpyxl
# Documentação: https://openpyxl.readthedocs.io/en/stable/ 

from pathlib import Path

from openpyxl import Workbook


ROOT_FOLDER = Path(__file__).parent
WORDBOOK_PATH = ROOT_FOLDER / 'workbook.xlsx'

workbook = Workbook()


sheet_name = "Minha planilha"
worksheet = workbook.create_sheet(sheet_name)
print(workbook.sheetnames)  # ['Sheet', 'Minha planilha']

workbook.remove(workbook['Sheet'])

#criando cabeçalhos
worksheet.cell(row=1, column=1, value='Name')
worksheet.cell(row=1, column=2, value='Age')
worksheet.cell(row=1, column=3, value='Note')


students = [
    #Name, Age, Note
    ['Maria', 22, 8.5  ],
    ['João', 20, 7.0  ],
    ['Ana',   21, 9.0  ],
    ['Pedro', 23, 6.5  ],
    ['Luiza', 22, 8.0  ],
]

for row_index, student in enumerate(students, start=2):
    worksheet.cell(row=row_index, column=1, value=student[0])
    worksheet.cell(row=row_index, column=2, value=student[1])
    worksheet.cell(row=row_index, column=3, value=student[2])

workbook.save(WORDBOOK_PATH)
print(f'workbook save in: {WORDBOOK_PATH}')