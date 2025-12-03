# openpyxl le e slterar dados de um aplanilha
# Com essa biblioteca será possível ler e escrever dados em células
# específicas, formatar células, inserir gráficos,
# criar fórmulas, adicionar imagens e outros elementos gráficos às suas
# planilhas. Ela é útil para automatizar tarefas envolvendo planilhas do
# Excel, como a criação de relatórios e análise de dados e/ou facilitando a
# manipulação de grandes quantidades de informações.
# Instalação necessária: pip install openpyxl
# Documentação: https://openpyxl.readthedocs.io/en/stable/ 

from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet

ROOT_FOLDER = Path(__file__).parent
WORDBOOK_PATH = ROOT_FOLDER / 'workbook.xlsx'

#carregando um arquivo do Excel
workbook: Workbook = load_workbook(WORDBOOK_PATH)

# nome da planilha
sheet_name = "Minha planilha"

# seleciona a planilha
worksheet: Worksheet = workbook[sheet_name]
print(workbook.sheetnames)  # ['Sheet', 'Minha planilha']

for row in worksheet.iter_rows(min_row=1, max_row=6, min_col=1, max_col=3):
    for cell in row:
        print(cell.value, end='\t')

        if cell.value == "Ana":
            worksheet.cell(cell.row, 2, 32)
    print() # Acessando célula específica


workbook.save(WORDBOOK_PATH)
# print(f'workbook save in: {WORDBOOK_PATH}')